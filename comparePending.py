import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Alignment
from copy import copy
import datetime
import sys
import glob
import os
import re


today_date = (datetime.datetime.today()).strftime('%Y-%m-%d')
yesterday_date = (datetime.datetime.today() - datetime.timedelta(days=1)).strftime('%Y-%m-%d')
data_found = False

xlsx_file_new = f"data-pending/NEW_OCZEKUJACE_{today_date}.xlsx"
xlsx_files_old = glob.glob("data-pending/OLD_OCZEKUJACE_*.xlsx")

# Regex pattern to extract the date part
pattern = re.compile(r"data-pending/OLD_OCZEKUJACE_(\d{4}-\d{2}-\d{2})\.xlsx")

# Dictionary to store file paths with extracted dates
files_with_dates = {}

for file_path in xlsx_files_old:
    match = pattern.search(file_path)
    if match:
        date_part = match.group(1)  # Extract the date in 'yyyy-mm-dd' format
        try:
            date_obj = datetime.datetime.strptime(date_part, "%Y-%m-%d")  # Convert to datetime object for sorting
            files_with_dates[file_path] = date_obj
        except ValueError:
            print(f"Skipping file with invalid date format: {file_path}")

# Sort files by date in descending order
sorted_files = sorted(files_with_dates.items(), key=lambda x: x[1], reverse=True)

# Keep the latest file and delete older ones
if sorted_files:
    xlsx_file_old = sorted_files[0][0]  # Keep the newest file
    files_to_delete = [file[0] for file in sorted_files[1:]]  # All except the newest

    for file_path in files_to_delete:
        try:
            os.remove(file_path)
            print(f"Deleted: {file_path}")
        except Exception as e:
            print(f"Error deleting file {file_path}: {e}")

    print(f"Latest file kept: {xlsx_file_old}")
else:
    print("No matching files found.")

# Load the Excel files
df2 = pd.read_excel(xlsx_file_new, engine="openpyxl")
try:
    df1 = pd.read_excel(xlsx_file_old, engine="openpyxl")
    new_entries = df2[~df2['ID'].isin(df1['ID'])]
except Exception as e:
    new_entries = df2
    print(f"Error reading XLSX file {xlsx_file_old}: {e}")

def compare_files():

    # # Merge DataFrames to find new entries
    # new_entries = df2[~df2['ID'].isin(df1['ID'])]

    # # Find rows in df2 that are not in df1
    # new_entries = pd.concat([df1, df2]).drop_duplicates(keep=False)
    new_entries = df2[~df2['ID'].isin(df1['ID'])]
    new_entries = new_entries.loc[:, ~new_entries.columns.str.contains('^Unnamed')]
    print(f"Number of new records after comparison: {len(new_entries)}")

    # Load the original workbook and select the active worksheet
    wb = load_workbook(xlsx_file_new)
    ws = wb.active

    # Create a new workbook for the new entries
    new_wb = Workbook()
    new_ws = new_wb.active

    # Copy column widths from the original worksheet
    for col_letter, col_dim in ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = col_dim.width

    # Create a NamedStyle for the header
    header_style = NamedStyle(name="header_style")
    header_cell = ws[1][0]  # First cell in the header row

    # Copy font from the original header
    header_style.font = copy(header_cell.font)

    # Copy fill from the original header
    header_style.fill = copy(header_cell.fill)

    # Copy border from the original header
    header_style.border = copy(header_cell.border)

    # Copy alignment from the original header
    header_style.alignment = copy(header_cell.alignment)

    # Append the NamedStyle to the new workbook
    new_wb.add_named_style(header_style)
    new_entries = new_entries.loc[:, ~new_entries.columns.str.contains('^Unnamed')]

    # Append new entries to the new worksheet
    for r_idx, row in enumerate(dataframe_to_rows(new_entries, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = new_ws.cell(row=r_idx, column=c_idx, value=value)
            # Apply header style to the first row
            if r_idx == 1:
                cell.style = header_style

    # Save the new entries to a new Excel file
    new_wb.save(f"data-pending/OCZEKUJACE_{today_date}.xlsx")

    # Check if there is any data beyond the header
    data_found = ws.max_row > 1
    
    # Return exit code (0 = False, 1 = True)
    sys.exit(1 if data_found else 0)

# Call the function and get the output
data_found_result = compare_files()