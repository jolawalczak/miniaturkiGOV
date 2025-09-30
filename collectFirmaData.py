import subprocess
import sys
from copy import copy
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import time
import json

# Get the page number from command-line arguments
if len(sys.argv) < 2:
    print("Usage: python collectFirmaData.py <filename>  [start_time]")
    sys.exit(1)

try:
    input_excel_file = sys.argv[1]
except ValueError:
    print("Error: Filename parameter must be a string type.")
    sys.exit(1)

# Extract optional body argument if provided
if len(sys.argv) > 2:
    start_time = float(sys.argv[2])
else:
    start_time = time.time()

output_excel_file = input_excel_file

id_column_name = "ID" # Column name that contains ID numbers
output_columns = ["WOJEWÓDZTWO", "ADRES KORESPONDENCYJNY", "TELEFON"]
batch_size = 25  # Number of records to process before writing to the file

# Load the workbook (this preserves formatting)
wb = load_workbook(input_excel_file)
ws = wb.active  # Assumes data is in the active worksheet

# Record original column widths
original_widths = {}
for col_letter in ws.column_dimensions:
    original_widths[col_letter] = ws.column_dimensions[col_letter].width

# Read the header (assumes header is in the first row)
header = [cell.value for cell in ws[1]]

# Check if the ID column exists
if id_column_name not in header:
    print(f"Error: Column '{id_column_name}' not found in the Excel file.")
    sys.exit(1)

# Find the column index (1-indexed) for ID
id_col_index = header.index(id_column_name) + 1

# Ensure output columns exist; add if missing while preserving formatting
column_indices = {}
template_cell = ws.cell(row=1, column=1)  # Take style from the first column header
for col_name in output_columns:
    if col_name in header:
        column_indices[col_name] = header.index(col_name) + 1
    else:
        new_col_index = len(header) + 1
        header.append(col_name)  # Add column to header list
        new_header_cell = ws.cell(row=1, column=new_col_index, value=col_name)

        # Copy formatting from the first header cell
        new_header_cell.font = copy(template_cell.font)
        new_header_cell.fill = copy(template_cell.fill)
        new_header_cell.border = copy(template_cell.border)
        new_header_cell.alignment = copy(template_cell.alignment)
        new_header_cell.number_format = template_cell.number_format

        column_indices[col_name] = new_col_index

# Find the last filled row to continue appending data
last_filled_row = ws.max_row

# Track consecutive failures
consecutive_failures = 0
MAX_FAILURES = 3  # Stop script if 3 consecutive IDs fail

# Function to get telephone number by running the external script
def get_firma_data(id_value):
    try:
        if id_value is None or str(id_value).strip() == "":
            return {col: "" for col in output_columns}  # Return empty values if ID is empty
        
        clean_id = str(id_value).replace(".0", "").strip()
        max_attempts = 2 # First attempt + one retry

        for attempt in range(max_attempts):

            time.sleep(4)
            current_time = datetime.now().strftime("%H:%M:%S")
            print(f"{current_time} - Fetching data for ID: {clean_id} (Attempt {attempt + 1})")

            result = subprocess.run(["python3", "getFirmaByID.py", clean_id],
                                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)

            stdout_output = result.stdout.strip()
            stderr_output = result.stderr.strip()

            if result.returncode == 0 and stderr_output == "":  # Successful execution, no errors
                # Check if the response is an actual telephone number (not an error message)
                if "Failed" in stdout_output or "Error" in stdout_output:
                    print(f"Detected error message in response: {stdout_output}")
                else:
                    try:
                        return json.loads(stdout_output)  # Parsowanie JSON-a przed zwróceniem
                    except json.JSONDecodeError:
                        print(f"Invalid JSON response for ID: {clean_id}")
                        return "__ERROR__"
            
            if result.returncode == 1 and ("Failed" in stdout_output or "Error" in stdout_output):
                # Check if the response is an actual telephone number (not an error message)
                print(f"Detected error message in response: {stdout_output}")

            if attempt == 0:  # First failure, wait before retrying
                print("Waiting 180 seconds before retrying...")
                time.sleep(180)
            else:  # Second failure, wait and move to the next ID
                print(f"Final attempt failed for ID: {clean_id}. Waiting 180 seconds before moving to next ID.")
                time.sleep(180)
                return "__ERROR__"
            
    except Exception as e:
        print(f"Error processing ID {id_value}: {e}")
        return "__ERROR__"
    
# Store batch of data before writing to the file
batch_data = []
row_counter = 0

# Process each row in the Excel file
for row in range(2, ws.max_row + 1):
    id_value = ws.cell(row=row, column=id_col_index).value
    firma_data = get_firma_data(id_value)

    if isinstance(firma_data, dict):
        # Parse JSON output
        try:
            if "firma" in firma_data and isinstance(firma_data["firma"], list) and len(firma_data["firma"]) > 0:
                firma = firma_data["firma"][0]  # Get first company
                adres = firma.get("adresKorespondencyjny", {})
                adres_formatted = f"{adres.get('ulica', '')} {adres.get('budynek', '')}/{adres.get('lokal', '')}, {adres.get('kod', '')} {adres.get('miasto', '')}".replace("/None", "").strip(", ")
                
                firma_data = {
                    "WOJEWÓDZTWO": adres.get("wojewodztwo", ""),
                    "ADRES KORESPONDENCYJNY": adres_formatted if adres else "",
                    "TELEFON": firma.get("telefon", ""),
                }
            else:
                firma_data = {
                    "WOJEWÓDZTWO": "",
                    "ADRES KORESPONDENCYJNY": "",
                    "TELEFON": "",
                }
        except json.JSONDecodeError:
            print(f"Invalid JSON response for ID: {id} and column: {id_column_name}")
            firma_data = {
                    "WOJEWÓDZTWO": "",
                    "ADRES KORESPONDENCYJNY": "",
                    "TELEFON": "",
            }
        batch_data.append((row, firma_data))
    else:
        batch_data.append((row, {col: "__ERROR__" for col in output_columns}))
    row_counter += 1

    if firma_data == "__ERROR__":
        firma_data = {
            "WOJEWÓDZTWO": "",
            "ADRES KORESPONDENCYJNY": "",
            "TELEFON": "",
        }
        consecutive_failures += 1
        print(f"Consecutive failures: {consecutive_failures}/{MAX_FAILURES}")

        if consecutive_failures >= MAX_FAILURES:
            print(f"❌ Stopping script. {MAX_FAILURES} consecutive IDs failed.")
            break  # Stop processing further IDs
    else:
        consecutive_failures = 0  # Reset failure counter if we get a successful result
        
        
    # Write batch to the Excel file after every 25 records
    if row_counter % batch_size == 0:
        print(f"💾 Saving {len(batch_data)} records to file after processing {row_counter} IDs...")
        for row_num, record in batch_data:
            if isinstance(record, dict):  # Ensure record is a dictionary
                for col_name, col_index in column_indices.items():
                    ws.cell(row=row_num, column=col_index, value=record.get(col_name, ""))
        wb.save(output_excel_file)
        batch_data.clear()
        print(f"✅ Batch saved. Processed {row_counter} records so far.")

        current_time = time.time()
        total_time = current_time - start_time
        print(f"Total execution time so far: {total_time:.2f} seconds")
        if (total_time > 36000):
            print(f"Total time exceeded maximum time 10 hours, it was in total: {total_time:.2f} seconds")
            break

# Write remaining records if there are any left
if batch_data:
    print(f"💾 Saving final batch of {len(batch_data)} records...")
    for row_num, record in batch_data:
        if isinstance(record, dict):
            for col_name, col_index in column_indices.items():
                ws.cell(row=row_num, column=col_index, value=record.get(col_name, ""))
    wb.save(output_excel_file)
    print(f"✅ Final batch of {len(batch_data)} records saved.")


# Adjust column widths dynamically
for col_name, col_index in column_indices.items():
    max_length = max(
        len(str(ws.cell(row=row, column=col_index).value))
        for row in range(1, ws.max_row + 1)
        if ws.cell(row=row, column=col_index).value
    )
    ws.column_dimensions[get_column_letter(col_index)].width = max_length + 2

wb.save(output_excel_file)
print(f"Processed data saved to {output_excel_file}")