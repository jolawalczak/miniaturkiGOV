import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

# Get the page number from command-line arguments
if len(sys.argv) < 2:
    print("Usage: python convert.py <filename>")
    sys.exit(1)

try:
    filename = sys.argv[1]
except ValueError:
    print("Error: Filename parameter must be a string type.")
    sys.exit(1)

# paths for CSV and XLSX files
csv_file = filename
xlsx_file = filename.replace(".csv", ".xlsx")

# f"data/OCZEKUJACE_{yesterday_date}.xlsx"

if csv_file is None or not os.path.exists(csv_file):
    print(f"Error: File {csv_file} does not exist.")
    exit(1)

# Define expected column headers
define_headers = [
    "ID", "NAZWA", "IMIĘ",
    "NAZWISKO", "NIP", "REGON", "DATA ROZPOCZĘCIA",
    "STATUS"
]

# Map CSV headers to match expected headers
column_mapping = {
    "id": "ID",
    "nazwa": "NAZWA",
    "wlasciciel_imie": "IMIĘ",
    "wlasciciel_nazwisko": "NAZWISKO",
    "wlasciciel_nip": "NIP",
    "wlasciciel_regon": "REGON",
    "dataRozpoczecia": "DATA ROZPOCZĘCIA",
    "status": "STATUS"
}

# Read CSV data and rename columns
df_csv = pd.read_csv(csv_file, encoding="utf-8-sig").rename(columns=column_mapping)

# Ensure all values are strings without '.0' or 'nan'
df_csv = df_csv.applymap(lambda x: "" if pd.isna(x) else str(x).replace(".0", "").strip())

# Ensure DataFrame is not empty after renaming
if df_csv.empty:
    print("Warning: CSV file is empty after renaming, nothing to write.")
    exit(0)

# Define formatting styles
header_font = Font(bold=True, color="000000")
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Flag to check if we are creating a new file
new_file_created = False

# Check if XLSX file exists
if os.path.exists(xlsx_file):
    # Load existing Excel file and append data
    book = load_workbook(xlsx_file)
    sheet = book.active
    start_row = sheet.max_row + 1  # Append below existing data
else:
    # Create new Excel file and add headers
    book = Workbook()
    sheet = book.active
    sheet.append(define_headers)  # Write headers
    
    # Apply styles to headers
    for col_num, header in enumerate(define_headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    start_row = 2  # Data starts below the headers
    new_file_created = True  # Mark that a new file was created

# Append new rows from CSV, ensuring all data remains as strings
for row in df_csv.astype(str).itertuples(index=False, name=None):
    sheet.append(row)

# Set column widths
display_widths = [10, 30, 15, 15, 15, 15, 15, 15, 80]
for col_num, width in enumerate(display_widths, start=1):
    sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

# Save the Excel file
book.save(xlsx_file)

# Print confirmation messages
if new_file_created:
    print(f"New Excel file {xlsx_file} created with headers.")
else:
    print(f"Data successfully added to {xlsx_file}")